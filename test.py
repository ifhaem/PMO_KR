from PyQt5 import QtWidgets, QtCore
from PyQt5.QtWidgets import QTableWidget, QTableWidgetItem, QMessageBox
from app2 import Ui_App
from pulp import *
import sys
import openpyxl, itertools

class myapp(QtWidgets.QMainWindow):
    def __init__(self):
        super(myapp, self).__init__()
        self.ui = Ui_App()
        self.ui.setupUi(self)
        self.ui.calcBtn.clicked.connect(self.calculateBtn)
        self.ui.fromExcel.clicked.connect(self.fromExcelBtn)
        self.ui.inExcel.clicked.connect(self.inExcelBtn)

    def calculateBtn(self):
        expX1 = self.ui.expX1.value()
        expX2 = self.ui.expX2.value()
        expX3 = self.ui.expX3.value()
        expX4 = self.ui.expX4.value()
        expX5 = self.ui.expX5.value()
        expY1 = self.ui.expY1.value()
        expY2 = self.ui.expY2.value()
        expY3 = self.ui.expY3.value()
        expY4 = self.ui.expY4.value()
        expY5 = self.ui.expY5.value()
        expZ1 = self.ui.expZ1.value()
        expZ2 = self.ui.expZ2.value()
        expZ3 = self.ui.expZ3.value()
        expZ4 = self.ui.expZ4.value()
        expZ5 = self.ui.expZ5.value()
        x1 = pulp.LpVariable("x1", 0, 10000 * expX1)
        x2 = pulp.LpVariable("x2", 0, 10000 * expX2)
        x3 = pulp.LpVariable("x3", 0, 10000 * expX3)
        x4 = pulp.LpVariable("x4", 0, 10000 * expX4)
        x5 = pulp.LpVariable("x5", 0, 10000 * expX5)
        y1 = pulp.LpVariable("y1", 0, 10000 * expY1)
        y2 = pulp.LpVariable("y2", 0, 10000 * expY2)
        y3 = pulp.LpVariable("y3", 0, 10000 * expY3)
        y4 = pulp.LpVariable("y4", 0, 10000 * expY4)
        y5 = pulp.LpVariable("y5", 0, 10000 * expY5)
        z1 = pulp.LpVariable("z1", 0, 10000 * expZ1)
        z2 = pulp.LpVariable("z2", 0, 10000 * expZ2)
        z3 = pulp.LpVariable("z3", 0, 10000 * expZ3)
        z4 = pulp.LpVariable("z4", 0, 10000 * expZ4)
        z5 = pulp.LpVariable("z5", 0, 10000 * expZ5)
        problem = pulp.LpProblem("0", sense=LpMinimize)
        problem += expX1 * x1 + expX2 * x2 + expX3 * x3 + expX4 * x4 + expX5 * x5 + expY1 * y1 + expY2 * y2 + expY3 * y3 + expY4 * y4 + expY5 * y5 + expZ1 * z1 + expZ2 * z2 + expZ3 * z3 + expZ4 * z4 + expZ5 * z5, "Функция цели"
        problem += x1 + x2 + x3 + x4 + x5 >= self.ui.minX.value(), "1"
        problem += y1 + y2 + y3 + y4 + y5 >= self.ui.minY.value(), "2"
        problem += z1 + z2 + z3 + z4 + z5 >= self.ui.minZ.value(), "3"
        problem += x1 + y1 + z1 <= self.ui.max1.value(), "4"
        problem += x2 + y2 + z2 <= self.ui.max2.value(), "5"
        problem += x3 + y3 + z3 <= self.ui.max3.value(), "6"
        problem += x4 + y4 + z4 <= self.ui.max4.value(), "7"
        problem += x5 + y5 + z5 <= self.ui.max5.value(), "8"
        problem.solve()
        t=0
        for i in range(0, 3):
            for j in range(0, 5):
                self.ui.result.setItem(i, j, QTableWidgetItem(str(problem.variables()[t].varValue)))
                t += 1
        self.ui.expensesSum.setText(str(value(problem.objective)))

    def fromExcelBtn(self):
        wb = openpyxl.load_workbook("KR2.xlsx")
        sheet = wb['Лист1']
        expX1 = sheet['B5'].value
        expX2 = sheet['C5'].value
        expX3 = sheet['D5'].value
        expX4 = sheet['E5'].value
        expX5 = sheet['F5'].value
        expY1 = sheet['B6'].value
        expY2 = sheet['C6'].value
        expY3 = sheet['D6'].value
        expY4 = sheet['E6'].value
        expY5 = sheet['F6'].value
        expZ1 = sheet['B7'].value
        expZ2 = sheet['C7'].value
        expZ3 = sheet['D7'].value
        expZ4 = sheet['E7'].value
        expZ5 = sheet['F7'].value
        expSet = [expX1, expX2, expX3, expX4, expX5, expY1, expY2, expY3, expY4, expY5, expZ1, expZ2, expZ3, expZ4, expZ5]
        self.ui.expX1.setValue(expX1)
        self.ui.expX2.setValue(expX2)
        self.ui.expX3.setValue(expX3)
        self.ui.expX4.setValue(expX4)
        self.ui.expX5.setValue(expX5)
        self.ui.expY1.setValue(expY1)
        self.ui.expY2.setValue(expY2)
        self.ui.expY3.setValue(expY3)
        self.ui.expY4.setValue(expY4)
        self.ui.expY5.setValue(expY5)
        self.ui.expZ1.setValue(expZ1)
        self.ui.expZ2.setValue(expZ2)
        self.ui.expZ3.setValue(expZ3)
        self.ui.expZ4.setValue(expZ4)
        self.ui.expZ5.setValue(expZ5)
        self.ui.minX.setValue(sheet['J11'].value)
        self.ui.minY.setValue(sheet['J12'].value)
        self.ui.minZ.setValue(sheet['J13'].value)
        self.ui.max1.setValue(sheet['B17'].value)
        self.ui.max2.setValue(sheet['C17'].value)
        self.ui.max3.setValue(sheet['D17'].value)
        self.ui.max4.setValue(sheet['E17'].value)
        self.ui.max5.setValue(sheet['F17'].value)
        self.ui.result.setItem(0, 0, QTableWidgetItem(str(sheet['B11'].value)))
        self.ui.result.setItem(0, 1, QTableWidgetItem(str(sheet['C11'].value)))
        self.ui.result.setItem(0, 2, QTableWidgetItem(str(sheet['D11'].value)))
        self.ui.result.setItem(0, 3, QTableWidgetItem(str(sheet['E11'].value)))
        self.ui.result.setItem(0, 4, QTableWidgetItem(str(sheet['F11'].value)))
        self.ui.result.setItem(1, 0, QTableWidgetItem(str(sheet['B12'].value)))
        self.ui.result.setItem(1, 1, QTableWidgetItem(str(sheet['C12'].value)))
        self.ui.result.setItem(1, 2, QTableWidgetItem(str(sheet['D12'].value)))
        self.ui.result.setItem(1, 3, QTableWidgetItem(str(sheet['E12'].value)))
        self.ui.result.setItem(1, 4, QTableWidgetItem(str(sheet['F12'].value)))
        self.ui.result.setItem(2, 0, QTableWidgetItem(str(sheet['B13'].value)))
        self.ui.result.setItem(2, 1, QTableWidgetItem(str(sheet['C13'].value)))
        self.ui.result.setItem(2, 2, QTableWidgetItem(str(sheet['D13'].value)))
        self.ui.result.setItem(2, 3, QTableWidgetItem(str(sheet['E13'].value)))
        self.ui.result.setItem(2, 4, QTableWidgetItem(str(sheet['F13'].value)))
        resultSet = [int(self.ui.result.item(0, 0).text()), int(self.ui.result.item(0, 1).text()), int(self.ui.result.item(0, 2).text()), int(self.ui.result.item(0, 3).text()), int(self.ui.result.item(0, 4).text()), int(self.ui.result.item(1, 0).text()), int(self.ui.result.item(1, 1).text()), int(self.ui.result.item(1, 2).text()), int(self.ui.result.item(1, 3).text()), int(self.ui.result.item(1, 4).text()), int(self.ui.result.item(2, 0).text()), int(self.ui.result.item(2, 1).text()), int(self.ui.result.item(2, 2).text()), int(self.ui.result.item(2, 3).text()), int(self.ui.result.item(2, 4).text())]
        expSum = 0
        for (a, b) in itertools.zip_longest(expSet, resultSet):
            expSum += a * b
        self.ui.expensesSum.setText(str(expSum))

    def inExcelBtn(self):
        msg = QMessageBox()
        msg.setWindowTitle("Сохранить")
        msg.setText("Сохранить в Excel?")
        msg.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
        if (msg.exec_() == QMessageBox.Ok):
            wb = openpyxl.load_workbook("KR2.xlsx")
            sheet = wb['Лист1']
            expX1 = self.ui.expX1.value()
            expX2 = self.ui.expX2.value()
            expX3 = self.ui.expX3.value()
            expX4 = self.ui.expX4.value()
            expX5 = self.ui.expX5.value()
            expY1 = self.ui.expY1.value()
            expY2 = self.ui.expY2.value()
            expY3 = self.ui.expY3.value()
            expY4 = self.ui.expY4.value()
            expY5 = self.ui.expY5.value()
            expZ1 = self.ui.expZ1.value()
            expZ2 = self.ui.expZ2.value()
            expZ3 = self.ui.expZ3.value()
            expZ4 = self.ui.expZ4.value()
            expZ5 = self.ui.expZ5.value()
            expSet = [expX1, expX2, expX3, expX4, expX5, expY1, expY2, expY3, expY4, expY5, expZ1, expZ2, expZ3, expZ4,
                      expZ5]
            x1 = pulp.LpVariable("x1", 0, 10000 * expX1)
            x2 = pulp.LpVariable("x2", 0, 10000 * expX2)
            x3 = pulp.LpVariable("x3", 0, 10000 * expX3)
            x4 = pulp.LpVariable("x4", 0, 10000 * expX4)
            x5 = pulp.LpVariable("x5", 0, 10000 * expX5)
            y1 = pulp.LpVariable("y1", 0, 10000 * expY1)
            y2 = pulp.LpVariable("y2", 0, 10000 * expY2)
            y3 = pulp.LpVariable("y3", 0, 10000 * expY3)
            y4 = pulp.LpVariable("y4", 0, 10000 * expY4)
            y5 = pulp.LpVariable("y5", 0, 10000 * expY5)
            z1 = pulp.LpVariable("z1", 0, 10000 * expZ1)
            z2 = pulp.LpVariable("z2", 0, 10000 * expZ2)
            z3 = pulp.LpVariable("z3", 0, 10000 * expZ3)
            z4 = pulp.LpVariable("z4", 0, 10000 * expZ4)
            z5 = pulp.LpVariable("z5", 0, 10000 * expZ5)
            problem = pulp.LpProblem("0", sense=LpMinimize)
            problem += expX1 * x1 + expX2 * x2 + expX3 * x3 + expX4 * x4 + expX5 * x5 + expY1 * y1 + expY2 * y2 + expY3 * y3 + expY4 * y4 + expY5 * y5 + expZ1 * z1 + expZ2 * z2 + expZ3 * z3 + expZ4 * z4 + expZ5 * z5, "Функция цели"
            problem += x1 + x2 + x3 + x4 + x5 >= self.ui.minX.value(), "1"
            problem += y1 + y2 + y3 + y4 + y5 >= self.ui.minY.value(), "2"
            problem += z1 + z2 + z3 + z4 + z5 >= self.ui.minZ.value(), "3"
            problem += x1 + y1 + z1 <= self.ui.max1.value(), "4"
            problem += x2 + y2 + z2 <= self.ui.max2.value(), "5"
            problem += x3 + y3 + z3 <= self.ui.max3.value(), "6"
            problem += x4 + y4 + z4 <= self.ui.max4.value(), "7"
            problem += x5 + y5 + z5 <= self.ui.max5.value(), "8"
            problem.solve()
            t = 0
            for i in range(5, 8):
                for j in range(2, 7):
                    sheet.cell(row=i, column=j).value = int(expSet[t])
                    t += 1
            sheet['J11'] = int(self.ui.minX.value())
            sheet['J12'] = int(self.ui.minY.value())
            sheet['J13'] = int(self.ui.minZ.value())
            sheet['B17'] = int(self.ui.max1.value())
            sheet['C17'] = int(self.ui.max2.value())
            sheet['D17'] = int(self.ui.max3.value())
            sheet['E17'] = int(self.ui.max4.value())
            sheet['F17'] = int(self.ui.max5.value())
            t = 0
            for i in range(11, 14):
                for j in range(2, 7):
                    sheet.cell(row=i, column=j).value = problem.variables()[t].varValue
                    t += 1
            sheet['J8'] = value(problem.objective)
            wb.save("KR2.xlsx")

app = QtWidgets.QApplication([])
application = myapp()
application.show()

sys.exit(app.exec())